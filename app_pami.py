"""
PAMI Internación — Procesador & Generador de Templates
=======================================================
App Streamlit para procesar archivos crudos INSSJP, valorizar
prácticas según atribución DIRECTA/NORMAL y generar templates
para importación en evweb.
"""

import os
import re
import shutil
import unicodedata
import zipfile
from datetime import datetime
from io import BytesIO

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import streamlit as st
from auth import login_required
login_required()


# ══════════════════════════════════════════════════════════════
#  CONFIGURACIÓN
# ══════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="PAMI Internación",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Estilos ──────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&display=swap');
:root {
    --azul-1: #0f2942;
    --azul-2: #1a3a5c;
    --azul-3: #2563eb;
    --gris-1: #f7f8fa;
    --gris-2: #e5e7eb;
    --verde:  #059669;
    --ambar:  #d97706;
}
.stApp { font-family: 'DM Sans', sans-serif; }

/* Header */
.hdr {
    background: linear-gradient(135deg, var(--azul-1), var(--azul-2));
    padding: 1.8rem 2.2rem; border-radius: 14px;
    margin-bottom: 1.6rem; color: #fff;
    box-shadow: 0 6px 24px rgba(15,41,66,.3);
}
.hdr h1 { margin:0; font-size:1.75rem; font-weight:700; letter-spacing:-.4px; }
.hdr p  { margin:.4rem 0 0; opacity:.8; font-size:.92rem; }

/* Stat cards */
.kpi { display:flex; gap:.8rem; margin-bottom:1.2rem; }
.kpi-card {
    flex:1; background:var(--gris-1); border:1px solid var(--gris-2);
    border-radius:10px; padding:.9rem 1.1rem; text-align:center;
}
.kpi-card .n { font-size:1.6rem; font-weight:700; color:var(--azul-1); }
.kpi-card .l { font-size:.7rem; color:#6b7280; text-transform:uppercase; letter-spacing:.8px; }

/* Banners */
.ok-box  { background:#ecfdf5; border:1px solid #6ee7b7; border-radius:10px; padding:.9rem 1.2rem; color:#065f46; margin:.6rem 0; }
.warn-box{ background:#fffbeb; border:1px solid #fcd34d; border-radius:10px; padding:.9rem 1.2rem; color:#92400e; margin:.6rem 0; }

/* Sidebar */
div[data-testid="stSidebar"] { background:linear-gradient(180deg,#f8fafc,#eef2f7); }

/* evweb button */
.ev-btn {
    display:inline-block; background:linear-gradient(135deg,#2563eb,#1d4ed8);
    color:#fff !important; padding:.65rem 1.3rem; border-radius:8px;
    text-decoration:none; font-weight:600; font-size:.9rem;
    box-shadow:0 4px 12px rgba(37,99,235,.3); transition:transform .15s;
}
.ev-btn:hover { transform:translateY(-2px); }

/* Step indicator */
.step-row { display:flex; gap:.5rem; margin-bottom:1rem; }
.step-pill {
    flex:1; text-align:center; padding:.45rem .6rem;
    border-radius:8px; font-size:.78rem; font-weight:600;
}
.step-done  { background:#d1fae5; color:#065f46; }
.step-active{ background:#dbeafe; color:#1e40af; border:2px solid #93c5fd; }
.step-todo  { background:#f3f4f6; color:#9ca3af; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
#  FUNCIONES DE PROCESAMIENTO PAMI
# ══════════════════════════════════════════════════════════════

def normalizar_nombre(nombre):
    """Normaliza nombres eliminando saltos de línea, espacios dobles, etc."""
    if pd.isna(nombre) or nombre is None:
        return ""
    n = str(nombre).upper().strip().replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", n).strip()


CODIGOS_DIRECTOS = {800001, 801001, 801002, 801003, 801004, 801005, 801008, 801009, 816005}


def es_directa_por_codigo(c_practica):
    try:
        return int(c_practica) in CODIGOS_DIRECTOS
    except (ValueError, TypeError):
        return False


def es_directa_por_texto(texto):
    """Verifica si un texto (prof o BATE) indica modalidad DIRECTA."""
    t = normalizar_nombre(texto)
    return any(x in t for x in ["RIOS PART", "RIOSPART", "DE SABATO", "DESABATO"])


def determinar_modalidad(row):
    if es_directa_por_codigo(row["C_PRACTICA"]):
        return "DIRECTA"
    if es_directa_por_texto(row.get("PROFESIONAL ACTUANTE", "")):
        return "DIRECTA"
    if es_directa_por_texto(row.get("BATE", "")):
        return "DIRECTA"
    return "NORMAL"


# ── Búsqueda en DIRECTA ─────────────────────────────────────
def buscar_en_directa(df_dir, bate, c_practica, prof_actuante):
    try:
        c_str = str(int(float(c_practica)))
    except (ValueError, TypeError):
        c_str = str(c_practica).strip()
    clave = str(bate).strip() + c_str

    for _, r in df_dir.iterrows():
        if str(r.get("Codigos de B", "")).strip() == clave:
            return r

    texto = normalizar_nombre(prof_actuante) or normalizar_nombre(bate)
    for _, r in df_dir.iterrows():
        bdb = normalizar_nombre(r["BATE"])
        cdb = normalizar_nombre(str(r.get("Codigos de B", "")))
        if any(x in texto for x in ["RIOS PART", "RIOSPART"]) and any(x in bdb for x in ["RIOS PART", "RIOSPART"]):
            return r
        if any(x in texto for x in ["DE SABATO", "DESABATO"]) and (
            any(x in bdb for x in ["DE SABATO", "DESABATO"]) or any(x in cdb for x in ["DE SABATO", "DESABATO"])
        ):
            return r
        if "RUFFINI" in texto and ("RUFFINI" in bdb or "RUFFINI" in cdb):
            return r
    return None


# ── Búsqueda en NORMAL ──────────────────────────────────────
def buscar_en_normal(df_nor, bate, prof_actuante, ayudante, anatomia):
    bate_n = normalizar_nombre(bate)
    prof_n = normalizar_nombre(prof_actuante)
    ay_s = str(ayudante).upper()
    an_s = str(anatomia).upper()
    clave = bate_n + prof_n + ay_s + an_s

    # 1) Exacta por Codigos de B
    for _, r in df_nor.iterrows():
        cb = normalizar_nombre(str(r.get("Codigos de B", "")))
        if cb == clave:
            return r

    # 2) Match de profesional normalizado
    for _, r in df_nor.iterrows():
        if normalizar_nombre(r["BATE"]) != bate_n:
            continue
        if str(r.get("Ayudante", "")).upper() != ay_s or str(r.get("Anatomia", "")).upper() != an_s:
            continue
        rp = normalizar_nombre(r.get("PROFESIONAL ACTUANTE", ""))
        if rp.startswith("CONDICION"):
            continue
        if rp == prof_n or rp in prof_n or prof_n in rp:
            return r
        pp = prof_n.replace("DR ", "").replace("DRA ", "").replace("DR. ", "").split()
        rps = rp.replace("DR ", "").replace("DRA ", "").replace("DR. ", "").split()
        if pp and rps and pp[0] == rps[0]:
            return r

    # 3) Fallback: CONDICION genérica
    for _, r in df_nor.iterrows():
        if normalizar_nombre(r["BATE"]) != bate_n:
            continue
        if str(r.get("Ayudante", "")).upper() != ay_s or str(r.get("Anatomia", "")).upper() != an_s:
            continue
        if normalizar_nombre(r.get("PROFESIONAL ACTUANTE", "")).startswith("CONDICION"):
            return r
    return None


# ── Valorización ─────────────────────────────────────────────
def valorizar_directa(row, match, monto):
    texto = normalizar_nombre(row.get("PROFESIONAL ACTUANTE", "")) or normalizar_nombre(row.get("BATE", ""))
    pct_hc = float(match.get("%HC", 0) or 0)
    pct_hp = float(match.get("%HP", 0) or 0)
    cuenta_hc = match.get("Cuenta HC", 0)
    cuenta_p = match.get("Cuenta P", 0)

    if any(x in texto for x in ["RIOS PART", "RIOSPART", "RUFFINI"]):
        pct, cta, tipo = (pct_hp, cuenta_p, "%HP") if pct_hp > 0 else (pct_hc, cuenta_hc, "%HC")
    elif any(x in texto for x in ["DE SABATO", "DESABATO"]):
        pct, cta, tipo = pct_hc, cuenta_hc, "%HC"
    else:
        pct, cta, tipo = pct_hc, cuenta_hc, "%HC"

    if pct > 0:
        f = row.copy()
        f["MONTO"] = round(monto * pct, 2)
        f["MATRICULA"] = cta
        f["MODALIDAD"] = "DIRECTA"
        f["TIPO_ATRIBUCION"] = tipo
        return [f]
    return []


def valorizar_normal(row, match, monto, cuenta_ayudante):
    configs = [
        ("%HC", "Cuenta HC"),
        ("%HP", "Cuenta P"),
        ("%Ay", None),
        ("%Ap", "Cuenta AP"),
        ("%G", "Cuenta G"),
    ]
    filas = []
    for pct_col, cta_col in configs:
        pct = float(match.get(pct_col, 0) or 0)
        if pct <= 0:
            continue
        if cta_col:
            cta = match.get(cta_col, 0)
            if not cta or str(cta) == "0":
                continue
        else:
            cta = cuenta_ayudante
            if not cta or not str(cta).strip():
                continue
        f = row.copy()
        f["MONTO"] = round(monto * pct, 2)
        f["MATRICULA"] = cta
        f["MODALIDAD"] = "NORMAL"
        f["TIPO_ATRIBUCION"] = pct_col
        filas.append(f)
    return filas


# ══════════════════════════════════════════════════════════════
#  GENERADOR DE TEMPLATES EVWEB
# ══════════════════════════════════════════════════════════════

SHEET_TEMPLATE = "ImportacionesEvweb"
FMT_NUM = '_-* #,##0.00_-;\\-* #,##0.00_-;_-* "-"??_-;_-@_-'
FMT_DATE = "m/d/yy h:mm"

COLS_TPL = [
    ("matricula",             7.89,  "General",  None,     None,     "General",  None,     None),
    ("afiliado_numero",       15.0,  "General",  "center", None,     "0",        None,     None),
    ("afiliado_denominacion", 41.66, "General",  "left",   None,     "General",  "left",   None),
    ("fecha_prestacion",      18.55, FMT_DATE,   "center", "center", "@",        "center", None),
    ("cod_practica",          11.66, "General",  None,     None,     "General",  None,     None),
    ("cantidad",              None,  "General",  None,     None,     "General",  None,     None),
    ("actuacion",             None,  "General",  None,     None,     "General",  None,     None),
    ("tipo_facturacion",      11.55, "General",  None,     None,     "General",  None,     None),
    ("honorario",             13.22, FMT_NUM,    None,     None,     "General",  None,     None),
    ("1er_ayudante",          11.55, "General",  None,     None,     "General",  None,     None),
    ("2do_ayudante",          None,  "General",  None,     None,     "General",  None,     None),
    ("gasto",                 11.66, "General",  None,     None,     "General",  None,     None),
    ("modulo",                11.55, "General",  None,     None,     "General",  None,     None),
    ("aparatoligia",          None,  "General",  None,     None,     "General",  None,     None),
    ("id_anticipo",           None,  "General",  None,     None,     "General",  None,     None),
    ("iva",                   11.66, "General",  "center", None,     "General",  None,     None),
    ("numaut",                15.33, "General",  "center", None,     "General",  None,     None),
    ("cuit",                  13.66, "General",  None,     None,     "General",  None,     None),
    ("fecha_presentacion",    16.78, FMT_DATE,   "center", "center", FMT_DATE,   "center", "center"),
    ("coseguro",              11.66, FMT_NUM,    None,     None,     "General",  None,     None),
    ("modalidadCoseguro",     11.55, "General",  None,     None,     "General",  None,     None),
    ("nroAutorizacion",       11.66, "General",  None,     None,     "General",  None,     None),
]

NOMBRES_TPL = [c[0] for c in COLS_TPL]
FONT_BASE = Font(name="MS Sans Serif", size=10, bold=False)
FONT_HDR = Font(name="MS Sans Serif", size=10, bold=True)

MAPEO_TPL = {
    "matricula":             ["cuenta", "cuenta_matricula", "matricula", "n° de cuenta"],
    "afiliado_numero":       ["credencial", "NRO. BENEFICIO/GP", "nro afiliado", "afiliado_numero"],
    "afiliado_denominacion": ["apellido y nombre", "afiliado_denominacion", "nom.beneficiario"],
    "fecha_prestacion":      ["F. DE PRACTICA", "fecha_transaccion", "fecha prestacion", "fecha"],
    "cod_practica":          ["C_PRACTICA", "cod_practica", "prestacion", "codigo"],
    "cantidad":              ["CANTIDAD INFORMADAS", "cantidad", "cant"],
    "honorario":             ["MONTO", "total", "honorario", "importe_total"],
    "iva":                   ["iva"],
    "numaut":                ["N. DE OP", "transaccion_item", "numaut", "NRO. ORDEN"],
    "coseguro":              ["copago", "coseguro"],
}

COLS_CERO_TPL = ["actuacion", "1er_ayudante", "2do_ayudante", "gasto", "modulo", "aparatoligia"]
COLS_VACIA_TPL = ["tipo_facturacion", "id_anticipo", "cuit", "modalidadCoseguro", "nroAutorizacion"]
FILAS_POR_TPL = 1500


def _quitar_acentos(t):
    return "".join(ch for ch in unicodedata.normalize("NFKD", t) if not unicodedata.combining(ch))


def _norm_col(t):
    if t is None:
        return ""
    return re.sub(r"[\s_\-\.]+", "", _quitar_acentos(str(t).strip().lower()))


def _buscar_col(cols_norm, candidatos):
    for c in candidatos:
        cn = _norm_col(c)
        for real, norm in cols_norm.items():
            if norm == cn:
                return real
    for c in candidatos:
        cn = _norm_col(c)
        for real, norm in cols_norm.items():
            if cn and cn in norm:
                return real
    return None


def _extraer_numero(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    d = re.sub(r"\D", "", str(v))
    return int(d) if d else None


def _ultimo_dia_mes_ant(hoy=None):
    hoy = hoy or datetime.now()
    return (hoy.replace(day=1) - pd.Timedelta(days=1)).replace(
        hour=0, minute=0, second=0, microsecond=0
    )


def crear_template_wb():
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_TEMPLATE
    for i, meta in enumerate(COLS_TPL, 1):
        nombre, ancho, _, _, _, h_nf, h_ha, h_va = meta
        if ancho:
            ws.column_dimensions[get_column_letter(i)].width = ancho
        c = ws.cell(row=1, column=i, value=nombre)
        c.font = FONT_HDR
        c.number_format = h_nf
        if h_ha or h_va:
            c.alignment = Alignment(horizontal=h_ha, vertical=h_va)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS_TPL))}1"
    return wb


def construir_filas_tpl(df):
    cols_n = {c: _norm_col(c) for c in df.columns}
    res = {d: _buscar_col(cols_n, cs) for d, cs in MAPEO_TPL.items()}
    fecha_pres = _ultimo_dia_mes_ant()
    filas = []
    for _, row in df.iterrows():
        f = {}
        col = res["matricula"]
        f["matricula"] = row[col] if col else None
        col = res["afiliado_numero"]
        f["afiliado_numero"] = _extraer_numero(row[col]) if col else None
        col = res["afiliado_denominacion"]
        f["afiliado_denominacion"] = row[col] if col else None
        col = res["fecha_prestacion"]
        val = row[col] if col else None
        if val is not None and not (isinstance(val, float) and pd.isna(val)):
            try:
                f["fecha_prestacion"] = pd.to_datetime(val).to_pydatetime()
            except Exception:
                f["fecha_prestacion"] = val
        else:
            f["fecha_prestacion"] = None
        col = res["cod_practica"]
        f["cod_practica"] = row[col] if col else None
        col = res["cantidad"]
        f["cantidad"] = row[col] if col else None
        for c in COLS_CERO_TPL:
            f[c] = 0
        for c in COLS_VACIA_TPL:
            f[c] = None
        col = res["honorario"]
        f["honorario"] = row[col] if col else None
        col = res["iva"]
        f["iva"] = row[col] if col else 0
        col = res["numaut"]
        f["numaut"] = _extraer_numero(row[col]) if col else None
        f["fecha_presentacion"] = fecha_pres
        col = res["coseguro"]
        f["coseguro"] = row[col] if col else 0
        filas.append(f)
    return filas


def generar_zip_templates(df, filas_por_tpl=FILAS_POR_TPL):
    filas = construir_filas_tpl(df)
    carpeta = "_tmp_tpl"
    os.makedirs(carpeta, exist_ok=True)
    rutas = []
    total = len(filas)
    n = max(1, (total + filas_por_tpl - 1) // filas_por_tpl) if total else 1
    for idx in range(n):
        ini, fin = idx * filas_por_tpl, (idx + 1) * filas_por_tpl
        wb = crear_template_wb()
        ws = wb.active
        for i, f in enumerate(filas[ini:fin], 2):
            for ci, nombre in enumerate(NOMBRES_TPL, 1):
                celda = ws.cell(row=i, column=ci, value=f.get(nombre))
                celda.font = FONT_BASE
                meta = COLS_TPL[ci - 1]
                celda.number_format = meta[2]
                if meta[3] or meta[4]:
                    celda.alignment = Alignment(horizontal=meta[3], vertical=meta[4])
        ruta = os.path.join(carpeta, f"Template_{idx + 1:02d}.xlsx")
        wb.save(ruta)
        wb.close()
        rutas.append(ruta)
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for r in rutas:
            zf.write(r, arcname=os.path.basename(r))
    shutil.rmtree(carpeta, ignore_errors=True)
    buf.seek(0)
    return buf, total, len(rutas)


def generar_excel_bytes(df):
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Valorizado")
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════
#  SESSION STATE
# ══════════════════════════════════════════════════════════════
for key, default in [
    ("paso", 0),
    ("df_proc", None),
    ("df_val", None),
    ("guardado", False),
    ("df_dir", None),
    ("df_nor", None),
    ("df_valores", None),
]:
    if key not in st.session_state:
        st.session_state[key] = default


# ══════════════════════════════════════════════════════════════
#  HEADER + SIDEBAR
# ══════════════════════════════════════════════════════════════
st.markdown("""
<div class="hdr">
    <h1>🏥 Procesador PAMI — Internación</h1>
    <p>Procesamiento, valorización y generación de templates para importación factura INSSJP</p>
</div>
""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("### 📁 Archivos de entrada")
    f_pami = st.file_uploader(
        "Archivo crudo PAMI", type=["xlsx", "xls"],
        help="Archivos extraidos desde la pagina del organismo para de internación. Es uno por cada facturacion efectuada",
    )
    f_base = st.file_uploader(
        "Base de Datos", type=["xlsx", "xls"],
        help="Parámetros de atribucion con solapas DIRECTA, NORMAL y VALORES",
    )
    st.markdown("---")
    st.markdown("### 🔗 Acceso rápido")
    st.markdown(
        '<a class="ev-btn" href="https://cmsc.evweb.com.ar/Account/Login?ReturnUrl=%2FfrmConceptos.aspx" '
        'target="_blank">🌐 Ir a evweb</a>',
        unsafe_allow_html=True,
    )

# ── Step indicator ───────────────────────────────────────────
step_labels = ["Cargar", "Editar", "Valorizar", "Exportar"]


def render_steps(activo):
    html = '<div class="step-row">'
    for i, lb in enumerate(step_labels):
        if i < activo:
            cls = "step-done"
            icon = "✓ "
        elif i == activo:
            cls = "step-active"
            icon = f"{i + 1}. "
        else:
            cls = "step-todo"
            icon = f"{i + 1}. "
        html += f'<div class="step-pill {cls}">{icon}{lb}</div>'
    html += "</div>"
    st.markdown(html, unsafe_allow_html=True)


render_steps(st.session_state.paso)


# ══════════════════════════════════════════════════════════════
#  CUERPO PRINCIPAL
# ══════════════════════════════════════════════════════════════
if f_pami and f_base:

    # ── Lectura de archivos ──────────────────────────────────
    try:
        xls_p = pd.ExcelFile(f_pami)
        hoja = None
        for s in xls_p.sheet_names:
            tmp = pd.read_excel(xls_p, sheet_name=s, nrows=2)
            if "MONTO" in tmp.columns and "C_PRACTICA" in tmp.columns:
                hoja = s
                break
        if not hoja:
            st.error("No se encontró hoja con columnas MONTO y C_PRACTICA en el archivo PAMI.")
            st.stop()
        df_pami = pd.read_excel(xls_p, sheet_name=hoja)
        df_pami["MONTO"] = pd.to_numeric(df_pami["MONTO"], errors="coerce").fillna(0)

        xls_b = pd.ExcelFile(f_base)
        if "DIRECTA" not in xls_b.sheet_names or "NORMAL" not in xls_b.sheet_names:
            st.error("La Base de Datos debe tener solapas DIRECTA y NORMAL.")
            st.stop()
        df_dir = pd.read_excel(xls_b, sheet_name="DIRECTA")
        df_nor = pd.read_excel(xls_b, sheet_name="NORMAL")
        df_val_sheet = (
            pd.read_excel(xls_b, sheet_name="VALORES")
            if "VALORES" in xls_b.sheet_names
            else None
        )

        st.session_state.df_dir = df_dir
        st.session_state.df_nor = df_nor
        st.session_state.df_valores = df_val_sheet
        st.session_state.paso = max(st.session_state.paso, 1)

    except Exception as e:
        st.error(f"Error al leer archivos: {e}")
        st.stop()

    # ── KPIs iniciales ───────────────────────────────────────
    filas_cero = int((df_pami["MONTO"] == 0).sum())
    st.markdown(f"""
    <div class="kpi">
        <div class="kpi-card"><div class="n">{len(df_pami)}</div><div class="l">Filas</div></div>
        <div class="kpi-card"><div class="n">${df_pami["MONTO"].sum():,.0f}</div><div class="l">Monto total</div></div>
        <div class="kpi-card"><div class="n">{filas_cero}</div><div class="l">Monto = 0</div></div>
        <div class="kpi-card"><div class="n">{df_pami["BATE"].nunique()}</div><div class="l">BATEs</div></div>
    </div>
    """, unsafe_allow_html=True)

    # ═════════════════════════════════════════════════════════
    #  PASO 1: COMPLETAR MONTOS Y PREPARAR EDICIÓN
    # ═════════════════════════════════════════════════════════
    st.markdown("### 🔧 Paso 1 — Procesar prácticas")

    # Completar MONTO = 0 desde VALORES
    if filas_cero > 0 and df_val_sheet is not None:
        # Detectar columnas dinámicamente (tolerante a variaciones)
        col_cprac_val = None
        col_valor_val = None
        for c in df_val_sheet.columns:
            cn = c.strip().upper().replace(" ", "_")
            if cn in ("C_PRACTICA", "CPRACTICA", "COD_PRACTICA", "CODIGO_PRACTICA"):
                col_cprac_val = c
            if "VALOR" in cn and "RESULT" in cn:
                col_valor_val = c
        # Fallback: buscar por contenido parcial
        if col_cprac_val is None:
            for c in df_val_sheet.columns:
                if "PRACT" in c.upper():
                    col_cprac_val = c
                    break
        if col_valor_val is None:
            for c in df_val_sheet.columns:
                if "VALOR" in c.upper():
                    col_valor_val = c
                    break

        if col_cprac_val and col_valor_val:
            # Convertir a numérico para comparación segura
            df_val_sheet[col_cprac_val] = pd.to_numeric(
                df_val_sheet[col_cprac_val], errors="coerce"
            )
            completados = 0
            for idx in df_pami[df_pami["MONTO"] == 0].index:
                try:
                    c_prac = int(df_pami.loc[idx, "C_PRACTICA"])
                except (ValueError, TypeError):
                    continue
                match = df_val_sheet[df_val_sheet[col_cprac_val] == c_prac]
                if len(match) > 0:
                    val = match.iloc[0][col_valor_val]
                    if pd.notna(val) and float(val) > 0:
                        df_pami.loc[idx, "MONTO"] = float(val)
                        completados += 1
            if completados:
                st.markdown(
                    f'<div class="ok-box">✅ Se completaron <b>{completados}</b> montos '
                    f"desde la solapa VALORES.</div>",
                    unsafe_allow_html=True,
                )
            restantes = int((df_pami["MONTO"] == 0).sum())
            if restantes:
                st.markdown(
                    f'<div class="warn-box">⚠️ Quedan <b>{restantes}</b> filas con '
                    f"MONTO = 0 sin valor en VALORES.</div>",
                    unsafe_allow_html=True,
                )
        else:
            st.markdown(
                '<div class="warn-box">⚠️ No se encontraron las columnas esperadas '
                'en la solapa VALORES (C_PRACTICA / VALOR RESULTANTE).</div>',
                unsafe_allow_html=True,
            )
    elif filas_cero > 0:
        st.markdown(
            f'<div class="warn-box">⚠️ {filas_cero} filas con MONTO = 0. '
            f"No se encontró solapa VALORES en la Base de Datos.</div>",
            unsafe_allow_html=True,
        )

    # Agregar columnas editables
    for col, default in [
        ("Ayudante", False),
        ("Cuenta_ayudante", ""),
        ("Anatomia", False),
        ("MATRICULA", ""),
    ]:
        if col not in df_pami.columns:
            df_pami[col] = default

    df_pami["_modalidad"] = df_pami.apply(determinar_modalidad, axis=1)

    # ── Filtros ──────────────────────────────────────────────
    st.markdown("#### 🔍 Filtros")
    fc1, fc2, fc3 = st.columns(3)
    with fc1:
        filtro_bate = st.selectbox(
            "BATE", ["Todos"] + sorted(df_pami["BATE"].unique().tolist())
        )
    with fc2:
        filtro_modal = st.selectbox("Modalidad", ["Todas", "DIRECTA", "NORMAL"])
    with fc3:
        filtro_monto = st.selectbox("Monto", ["Todos", "Solo $0", "Solo > $0"])

    df_vista = df_pami.copy()
    if filtro_bate != "Todos":
        df_vista = df_vista[df_vista["BATE"] == filtro_bate]
    if filtro_modal != "Todas":
        df_vista = df_vista[df_vista["_modalidad"] == filtro_modal]
    if filtro_monto == "Solo $0":
        df_vista = df_vista[df_vista["MONTO"] == 0]
    elif filtro_monto == "Solo > $0":
        df_vista = df_vista[df_vista["MONTO"] > 0]

    # ── Editor de datos ──────────────────────────────────────
    st.markdown("#### ✏️ Edición de prácticas")
    st.caption(
        "Modifique **Ayudante**, **Cuenta ayudante**, **Anatomía** y **MATRICULA**. "
        "Las demás columnas son de solo lectura."
    )

    cols_show = [
        "BATE", "PROFESIONAL ACTUANTE", "C_PRACTICA", "PRACTICA",
        "APELLIDO Y NOMBRE", "MONTO", "_modalidad",
        "Ayudante", "Cuenta_ayudante", "Anatomia", "MATRICULA",
    ]
    cols_show = [c for c in cols_show if c in df_vista.columns]

    col_cfg = {
        "Ayudante": st.column_config.CheckboxColumn("Ayudante", default=False),
        "Anatomia": st.column_config.CheckboxColumn("Anatomía", default=False),
        "Cuenta_ayudante": st.column_config.TextColumn("Cuenta Ayudante"),
        "MATRICULA": st.column_config.TextColumn("Matrícula"),
        "MONTO": st.column_config.NumberColumn("Monto", format="$%.2f"),
        "_modalidad": st.column_config.TextColumn("Modalidad", disabled=True),
        "BATE": st.column_config.TextColumn("BATE", disabled=True),
        "PROFESIONAL ACTUANTE": st.column_config.TextColumn("Prof. Actuante", disabled=True),
        "C_PRACTICA": st.column_config.TextColumn("Cód. Práctica", disabled=True),
        "PRACTICA": st.column_config.TextColumn("Práctica", disabled=True, width="large"),
        "APELLIDO Y NOMBRE": st.column_config.TextColumn("Paciente", disabled=True),
    }

    df_edit = st.data_editor(
        df_vista[cols_show],
        column_config=col_cfg,
        use_container_width=True,
        num_rows="fixed",
        key="editor_p1",
        height=420,
    )

    # Alerta de ayudante sin cuenta
    if df_edit is not None:
        sin_cta = df_edit[
            (df_edit["Ayudante"] == True)
            & (
                df_edit["Cuenta_ayudante"]
                .astype(str)
                .str.strip()
                .isin(["", "nan", "None"])
            )
        ]
        if len(sin_cta) > 0:
            st.markdown(
                f'<div class="warn-box">🚨 <b>{len(sin_cta)}</b> fila(s) con '
                f"Ayudante = TRUE sin Cuenta Ayudante. "
                f"Completar antes de valorizar.</div>",
                unsafe_allow_html=True,
            )

    # ── Vista detallada ──────────────────────────────────────
    with st.expander("📋 Vista detallada de una fila"):
        fila_n = st.number_input("Nro. fila:", 0, max(0, len(df_edit) - 1), 0)
        if len(df_edit) > 0:
            fd = df_edit.iloc[fila_n]
            c1, c2 = st.columns(2)
            with c1:
                for k in ["BATE", "PROFESIONAL ACTUANTE", "PRACTICA", "C_PRACTICA"]:
                    if k in fd.index:
                        st.markdown(f"**{k}:** {fd[k]}")
            with c2:
                for k in [
                    "APELLIDO Y NOMBRE", "MONTO", "_modalidad",
                    "Ayudante", "Anatomia", "Cuenta_ayudante", "MATRICULA",
                ]:
                    if k in fd.index:
                        st.markdown(f"**{k}:** {fd[k]}")

    # ── Guardar cambios ──────────────────────────────────────
    st.markdown("---")
    if st.button("💾 Guardar cambios y continuar", type="primary", use_container_width=True):
        for col in ["Ayudante", "Cuenta_ayudante", "Anatomia", "MATRICULA", "MONTO"]:
            if col in df_edit.columns:
                df_pami.loc[df_vista.index, col] = df_edit[col].values
        df_pami["_modalidad"] = df_pami.apply(determinar_modalidad, axis=1)
        st.session_state.df_proc = df_pami.copy()
        st.session_state.guardado = True
        st.session_state.paso = max(st.session_state.paso, 2)
        st.rerun()

    # ═════════════════════════════════════════════════════════
    #  PASO 2-3: VALORIZACIÓN
    # ═════════════════════════════════════════════════════════
    if st.session_state.guardado and st.session_state.df_proc is not None:
        st.markdown("---")
        st.markdown("### 💰 Paso 2-3 — Valorización")

        if st.button("🚀 Ejecutar valorización", type="primary", use_container_width=True):
            df_p = st.session_state.df_proc.copy()
            d_dir = st.session_state.df_dir
            d_nor = st.session_state.df_nor
            todas, errores = [], []
            bar = st.progress(0, text="Valorizando…")

            for i, (idx, row) in enumerate(df_p.iterrows()):
                bar.progress((i + 1) / len(df_p))
                monto = float(row.get("MONTO", 0) or 0)
                mod = row["_modalidad"]

                if monto == 0:
                    f = row.copy()
                    f["MODALIDAD"] = mod
                    f["TIPO_ATRIBUCION"] = "SIN MONTO"
                    todas.append(f)
                    continue

                if mod == "DIRECTA":
                    m = buscar_en_directa(
                        d_dir, row["BATE"], row["C_PRACTICA"],
                        row.get("PROFESIONAL ACTUANTE", ""),
                    )
                    if m is not None:
                        v = valorizar_directa(row, m, monto)
                        if v:
                            todas.extend(v)
                        else:
                            f = row.copy()
                            f["MODALIDAD"] = "DIRECTA"
                            f["TIPO_ATRIBUCION"] = "SIN MATCH %"
                            todas.append(f)
                            errores.append(f"#{idx} DIRECTA sin % > 0: {row['BATE']}")
                    else:
                        f = row.copy()
                        f["MODALIDAD"] = "DIRECTA"
                        f["TIPO_ATRIBUCION"] = "NO ENCONTRADO"
                        todas.append(f)
                        errores.append(
                            f"#{idx} No encontrado en DIRECTA: "
                            f"{row['BATE']}+{row['C_PRACTICA']}"
                        )
                else:  # NORMAL
                    ay = str(row.get("Ayudante", False)).upper()
                    an = str(row.get("Anatomia", False)).upper()
                    ca = str(row.get("Cuenta_ayudante", "")).strip()
                    m = buscar_en_normal(
                        d_nor, row["BATE"],
                        row.get("PROFESIONAL ACTUANTE", ""), ay, an,
                    )
                    if m is not None:
                        v = valorizar_normal(row, m, monto, ca)
                        if v:
                            todas.extend(v)
                        else:
                            f = row.copy()
                            f["MODALIDAD"] = "NORMAL"
                            f["TIPO_ATRIBUCION"] = "SIN MATCH %"
                            todas.append(f)
                            errores.append(f"#{idx} NORMAL sin % > 0: {row['BATE']}")
                    else:
                        f = row.copy()
                        f["MODALIDAD"] = "NORMAL"
                        f["TIPO_ATRIBUCION"] = "NO ENCONTRADO"
                        todas.append(f)
                        errores.append(
                            f"#{idx} No encontrado en NORMAL: "
                            f"{row['BATE']}+{row.get('PROFESIONAL ACTUANTE', '')}"
                        )

            bar.empty()

            if todas:
                df_val = pd.DataFrame(todas).drop(
                    columns=["_modalidad"], errors="ignore"
                )
                st.session_state.df_val = df_val
                st.session_state.paso = 3
                st.markdown(
                    f'<div class="ok-box">✅ Valorización finalizada: '
                    f"<b>{len(df_val)}</b> filas generadas a partir de "
                    f"{len(df_p)} originales.</div>",
                    unsafe_allow_html=True,
                )
                if errores:
                    with st.expander(f"⚠️ {len(errores)} advertencias"):
                        for e in errores:
                            st.warning(e)
            else:
                st.error("No se generaron filas valorizadas.")

    # ═════════════════════════════════════════════════════════
    #  PASO 4: RESULTADOS Y EXPORTACIÓN
    # ═════════════════════════════════════════════════════════
    if st.session_state.df_val is not None:
        st.markdown("---")
        st.markdown("### 📤 Paso 4 — Resultados y exportación")
        df_v = st.session_state.df_val

        directas = (
            int((df_v["MODALIDAD"] == "DIRECTA").sum())
            if "MODALIDAD" in df_v.columns
            else 0
        )
        normales = (
            int((df_v["MODALIDAD"] == "NORMAL").sum())
            if "MODALIDAD" in df_v.columns
            else 0
        )
        total_m = df_v["MONTO"].sum() if "MONTO" in df_v.columns else 0

        st.markdown(f"""
        <div class="kpi">
            <div class="kpi-card"><div class="n">{len(df_v)}</div><div class="l">Filas generadas</div></div>
            <div class="kpi-card"><div class="n">{directas}</div><div class="l">Directas</div></div>
            <div class="kpi-card"><div class="n">{normales}</div><div class="l">Normales</div></div>
            <div class="kpi-card"><div class="n">${total_m:,.0f}</div><div class="l">Monto total</div></div>
        </div>
        """, unsafe_allow_html=True)

        tab1, tab2, tab3 = st.tabs([
            "📋 Tabla completa", "🔍 Por modalidad", "📊 Resumen por BATE",
        ])
        with tab1:
            st.dataframe(df_v, use_container_width=True, height=380)
        with tab2:
            fm = st.selectbox(
                "Filtrar:",
                ["TODAS", "DIRECTA", "NORMAL", "NO ENCONTRADO", "SIN MONTO"],
                key="filt_mod",
            )
            dfx = (
                df_v
                if fm == "TODAS"
                else df_v[df_v["MODALIDAD"] == fm]
                if "MODALIDAD" in df_v.columns
                else df_v
            )
            st.dataframe(dfx, use_container_width=True, height=350)
        with tab3:
            if "BATE" in df_v.columns:
                res = (
                    df_v.groupby("BATE")
                    .agg(Filas=("MONTO", "count"), Monto_Total=("MONTO", "sum"))
                    .reset_index()
                )
                res["Monto_Total"] = res["Monto_Total"].round(2)
                st.dataframe(res, use_container_width=True)

        # ── Descargas ────────────────────────────────────────
        st.markdown("---")
        st.markdown("#### 📥 Descargas")
        fecha_str = datetime.now().strftime("%Y%m%d")
        c1, c2, c3 = st.columns(3)

        with c1:
            st.download_button(
                "📥 Excel valorizado",
                generar_excel_bytes(df_v),
                f"PAMI_Valorizado_{fecha_str}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )

        with c2:
            if st.button("📦 Generar Templates evweb", use_container_width=True):
                with st.spinner("Generando templates…"):
                    buf_zip, n_filas, n_tpl = generar_zip_templates(df_v)
                st.session_state["_zip_buf"] = buf_zip
                st.session_state["_zip_info"] = (n_filas, n_tpl)
                st.rerun()

            if "_zip_buf" in st.session_state:
                n_f, n_t = st.session_state["_zip_info"]
                st.markdown(
                    f'<div class="ok-box">✅ {n_f} filas → {n_t} template(s)</div>',
                    unsafe_allow_html=True,
                )
                st.download_button(
                    "⬇️ Descargar Templates.zip",
                    st.session_state["_zip_buf"],
                    f"Templates_{fecha_str}.zip",
                    "application/zip",
                    use_container_width=True,
                )

        with c3:
            st.markdown(
                '<a class="ev-btn" href="https://cmsc.evweb.com.ar/Account/Login'
                '?ReturnUrl=%2FfrmConceptos.aspx" target="_blank" '
                'style="display:block;text-align:center;margin-top:.3rem;">'
                "🌐 Abrir evweb</a>",
                unsafe_allow_html=True,
            )

else:
    # ── Pantalla de bienvenida ───────────────────────────────
    st.markdown("""
    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:14px;
                padding:2rem 2.5rem;margin-top:.5rem;">
        <h3 style="color:#0f2942;margin-top:0;">👋 Bienvenido</h3>
        <p>Para comenzar, cargue los dos archivos en la barra lateral:</p>
        <ol>
            <li><b>Archivo crudo PAMI</b> — factura de internación INSSJP</li>
            <li><b>Base de Datos</b> — parámetros con solapas DIRECTA, NORMAL y VALORES</li>
        </ol>
        <p style="margin-bottom:0;">El sistema procesará las prácticas, permitirá editarlas,
        valorizará según atribución DIRECTA/NORMAL y generará los templates para evweb.</p>
    </div>
    """, unsafe_allow_html=True)

# ── Footer ───────────────────────────────────────────────────
st.markdown("---")
st.markdown(
    '<p style="text-align:center;color:#94a3b8;font-size:.75rem;">'
    "Procesador PAMI Internación v2.0 · INSSJP · "
    "Colegio Médico del Sur del Chubut</p>",
    unsafe_allow_html=True,
)
